from pathlib import Path

from openpyxl import load_workbook

from django.core.management.base import BaseCommand

from apps.solicitacoes.models import Municipio


class Command(BaseCommand):
    help = "Importa municípios da Bahia"

    def handle(self, *args, **options):

        arquivo = Path("Municipios_Bahia_Preenchido.xlsx")

        if not arquivo.exists():
            self.stdout.write(
                self.style.ERROR(
                    f"Arquivo não encontrado: {arquivo}"
                )
            )
            return

        wb = load_workbook(arquivo)

        ws = wb.active

        total = 0

        for linha in ws.iter_rows(min_row=2, values_only=True):

            nome = linha[0]

            if not nome:
                continue

            Municipio.objects.get_or_create(
                nome=nome.strip()
            )

            total += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"{total} municípios importados com sucesso!"
            )
        )